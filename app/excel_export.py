# app/excel_export.py
# Exports DB data to Excel files in the exact format make_cv expects.
# Column names must exactly match what make_cv reads — verified from source.

import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime


def _safe_year(val):
    """Extract integer year from datetime, date, string, or int."""
    if val is None:
        return None
    if hasattr(val, 'year'):
        return int(val.year)
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _safe_date(val):
    """Return date object or None — strip time component if datetime."""
    if val is None:
        return None
    if hasattr(val, 'date'):
        return val.date()
    return val


def _safe_float(val):
    """Convert Decimal/None to float safely."""
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _wb(headers):
    """Create a workbook with a Data sheet and given headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(headers)
    return wb, ws


def write_personal_awards(path, rows):
    """
    PERSONALAWARDS joined with AWARDS
    make_cv expects: Title, Type, Year
    """
    wb, ws = _wb(['Title', 'Type', 'Year'])
    for r in rows:
        ws.append([
            r.get('Title', ''),
            r.get('Award Type', ''),
            _safe_year(r.get('Year')),
        ])
    wb.save(path)


def write_student_awards(path, rows):
    """
    STUDENTAWARDS joined with AWARDS
    make_cv expects: Student, Title, Amount, Category, Type, Year
    """
    wb, ws = _wb(['Student', 'Title', 'Amount', 'Category', 'Type', 'Year'])
    for r in rows:
        ws.append([
            r.get('Student', ''),
            r.get('Title', ''),
            r.get('Amount', 0),
            r.get('Category', ''),
            r.get('Award Type', ''),
            _safe_year(r.get('Year')),
        ])
    wb.save(path)


def write_proposals_and_grants(path, rows):
    """
    PROPOSAL + GRANTS combined
    make_cv expects: Sponsor, Allocated Amt, Total Cost, Funded?,
                     Title, Begin Date, End Date, Submit Date,
                     Principal Investigators
    """
    wb, ws = _wb([
        'Sponsor', 'Allocated Amt', 'Total Cost', 'Funded?',
        'Title', 'Begin Date', 'End Date', 'Submit Date',
        'Principal Investigators'
    ])
    for r in rows:
        funded = 'Y' if r.get('Funded?') in (1, True, 'Y', 'y', '1') else 'N'
        ws.append([
            r.get('Sponsor', ''),
            r.get('Allocated Amount', 0) or 0,
            r.get('Total Cost', 0) or 0,
            funded,
            r.get('Title', ''),
            _safe_date(r.get('Begin Date')),
            _safe_date(r.get('End Date')),
            _safe_date(r.get('Submit Date')),
            r.get('Principal Investigator', '') or r.get('Principal Investigators', ''),
        ])
    wb.save(path)


def write_grants(path, rows):
    """
    GRANTS only (funded only)
    make_cv expects: Sponsor, Allocated Amt, Total Cost, Funded?,
                     Title, Begin Date, End Date, Submit Date,
                     Principal Investigators
    """
    funded_rows = [r for r in rows if r.get('Funded?') in (1, True, 'Y', 'y', '1')]
    write_proposals_and_grants(path, funded_rows)


def write_expenditures(path, rows):
    """
    EXPENDITURE
    make_cv reads last row: Year, Name, Expenditure, Indirect, Tuition, Salary Recovery
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(['Year', 'Name', 'Expenditure', 'Indirect', 'Tuition', 'Salary Recovery'])
    for r in rows:
        ws.append([
            _safe_year(r.get('Year')),
            r.get('Name', ''),
            r.get('Expenditure', 0) or 0,
            r.get('Indirect', 0) or 0,
            r.get('Tuition', 0) or 0,
            r.get('Salary Recovery', 0) or 0,
        ])
    wb.save(path)


def write_current_students(path, rows):
    """
    CURRENTSTUDENTS
    make_cv expects: Student Name, Current Program, Start Date
    """
    wb, ws = _wb(['Student Name', 'Current Program', 'Start Date'])
    for r in rows:
        ws.append([
            r.get('Student Name', ''),
            r.get('Current Program', ''),
            _safe_date(r.get('Start Date')),
        ])
    wb.save(path)


def write_thesis(path, rows):
    """
    THESIS
    make_cv expects: Student, Start Date, Year, Degree, Advisor, Title, Comments
    Verified against Bohl Douglas real data.
    """
    wb, ws = _wb(['Student', 'Start Date', 'Year', 'Degree', 'Advisor', 'Title', 'Comments'])
    for r in rows:
        # make_cv reads Start Date as int — use 0 as placeholder (it reads year from Year column)
        start = r.get('Start Date')
        start_int = int(start.year) if hasattr(start, 'year') else (int(start) if start else 0)
        ws.append([
            r.get('Student Name', ''),
            start_int,
            _safe_year(r.get('Year')) or 0,
            r.get('Degree', ''),
            r.get('Advisor', ''),
            r.get('Title', ''),
            r.get('Comments', ''),
        ])
    wb.save(path)


def write_service(path, rows):
    """
    SERVICE
    make_cv expects: Description, Type, Position, Term, Calendar Year,
                     Hours/Semester, Comments
    """
    wb, ws = _wb([
        'Description', 'Type', 'Position', 'Term',
        'Calendar Year', 'Hours/Semester', 'Comments'
    ])
    for r in rows:
        hours = r.get('Hours/Semester')
        # Strip newlines from Description — they break LaTeX table rows
        desc = (r.get('Description', '') or '').replace('\n', ' ').replace('\r', ' ').strip()
        comments = (r.get('Comments', '') or '').replace('\n', ' ').replace('\r', ' ').strip()
        ws.append([
            desc,
            r.get('Type', ''),
            r.get('Position', ''),
            r.get('Term', ''),
            _safe_year(r.get('Calendar Year')),
            float(hours) if hours is not None else None,
            comments,
        ])
    wb.save(path)


def write_reviews(path, rows):
    """
    REVIEWS
    make_cv expects: Journal, Start, Rounds
    """
    wb, ws = _wb(['Journal', 'Start', 'Rounds'])
    for r in rows:
        ws.append([
            r.get('Journal', ''),
            _safe_date(r.get('Start Date')),
            r.get('Rounds', 0) or 0,
        ])
    wb.save(path)


def write_professional_development(path, rows):
    """
    PROFESSIONALDEVELOPMENT
    make_cv expects: Description, Type, Term, Calendar Year, Hours, Notes
    """
    wb, ws = _wb(['Description', 'Type', 'Term', 'Calendar Year', 'Hours', 'Notes'])
    for r in rows:
        ws.append([
            r.get('Description', ''),
            r.get('Type', ''),
            r.get('Term', ''),
            _safe_year(r.get('Calendar Year')),
            r.get('Hours', 0) or 0,
            r.get('Notes', ''),
        ])
    wb.save(path)


def write_undergraduate_research(path, rows):
    """
    UNDERGRADUATERESEARCH
    make_cv expects: Students, Title, Program Type, Term, Calendar Year
    """
    wb, ws = _wb(['Students', 'Title', 'Program Type', 'Term', 'Calendar Year'])
    for r in rows:
        ws.append([
            r.get('Students', ''),
            r.get('Title', ''),
            r.get('Program Type', ''),
            r.get('Term', ''),
            _safe_year(r.get('Calendar Year')) or r.get('Calendar Year'),
        ])
    wb.save(path)


def write_advisee_counts(path, rows):
    """
    ADVISEECOUNT
    make_far reads last row: Advisor Name, Count Distinct Name, YEAR
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(['Advisor Name', 'Count Distinct Name', 'YEAR'])
    for r in rows:
        ws.append([
            r.get('Advisor Name', ''),
            r.get('Advisee Count', 0) or 0,
            _safe_year(r.get('Year')),
        ])
    wb.save(path)


def write_advising_evaluation(path, rows):
    """
    ADVISINGEVALUATION — pass through all columns as-is
    make_cv reads this file but exact columns vary; write what we have.
    """
    if not rows:
        return
    headers = list(rows[0].keys())
    wb, ws = _wb(headers)
    for r in rows:
        ws.append([r.get(h, '') for h in headers])
    wb.save(path)


def write_teaching_evaluation(path, rows):
    """
    TEACHINGEVALUATION
    Installed make_cv (0.9.7) expects exactly these columns in Data sheet:
    term, combined_course_num, combined_num_sec, course_title,
    enrollment, count_19, mean_19, count_20, mean_20

    DB column mapping:
    Term            → term
    Combined Course Number → combined_course_num
    Course Section  → combined_num_sec (used as section identifier)
    Course Title    → course_title
    Enrolment       → enrollment
    Count Evals     → count_19 (number of respondents)
    Calculated Mean → mean_19  (overall mean score Q19)
    Count Evals     → count_20 (same respondents for Q20)
    Weighted Average→ mean_20  (weighted average score Q20)
    """
    wb, ws = _wb([
        'term', 'combined_course_num', 'combined_num_sec',
        'course_title', 'enrollment',
        'count_19', 'mean_19', 'count_20', 'mean_20'
    ])
    for r in rows:
        count = int(r.get('Count Evals') or 0)
        mean_19 = float(r.get('Calculated Mean') or 0)
        mean_20 = float(r.get('Weighted Average') or 0)
        enrol = int(r.get('Enrolment') or 0)
        ws.append([
            str(r.get('Term', '')),
            str(r.get('Combined Course Number', '') or ''),
            str(r.get('Course Section', '') or ''),
            str(r.get('Course Title', '') or ''),
            enrol,
            count,
            mean_19,
            count,
            mean_20,
        ])
    wb.save(path)


def write_prospective_visits(path, rows):
    """
    PROSPECTIVEVISIT
    make_far reads: Visits, Deposits, Year from last row
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(['Year', 'Staff', 'Visits', 'Deposits'])
    for r in rows:
        ws.append([
            _safe_year(r.get('Year')),
            r.get('Staff', ''),
            r.get('Visits', 0) or 0,
            r.get('Deposits', 0) or 0,
        ])
    wb.save(path)


def export_all(professor_key, professor_folder, db_data):
    """
    Write all Excel files for a professor into the correct subfolders.

    professor_folder: path like /path/to/Departments/Dept/Last, First
    db_data: dict with keys matching table names, values are lists of row dicts
    """
    def ensure(subfolder):
        p = os.path.join(professor_folder, subfolder)
        os.makedirs(p, exist_ok=True)
        return p

    awards_dir  = ensure('Awards')
    grants_dir  = ensure('Proposals & Grants')
    scholar_dir = ensure('Scholarship')
    service_dir = ensure('Service')
    teach_dir   = ensure('Teaching')

    write_personal_awards(
        os.path.join(awards_dir, 'personal awards data.xlsx'),
        db_data.get('personal_awards', [])
    )
    write_student_awards(
        os.path.join(awards_dir, 'student awards data.xlsx'),
        db_data.get('student_awards', [])
    )
    write_proposals_and_grants(
        os.path.join(grants_dir, 'proposals & grants.xlsx'),
        db_data.get('proposals', [])
    )
    write_grants(
        os.path.join(grants_dir, 'grants.xlsx'),
        db_data.get('grants', [])
    )
    write_expenditures(
        os.path.join(grants_dir, 'expenditures.xlsx'),
        db_data.get('expenditures', [])
    )
    write_current_students(
        os.path.join(scholar_dir, 'current student data.xlsx'),
        db_data.get('current_students', [])
    )
    write_thesis(
        os.path.join(scholar_dir, 'thesis data.xlsx'),
        db_data.get('thesis', [])
    )
    write_service(
        os.path.join(service_dir, 'service data.xlsx'),
        db_data.get('service', [])
    )
    write_reviews(
        os.path.join(service_dir, 'reviews data.xlsx'),
        db_data.get('reviews', [])
    )
    write_professional_development(
        os.path.join(service_dir, 'professional development data.xlsx'),
        db_data.get('prof_development', [])
    )
    write_undergraduate_research(
        os.path.join(service_dir, 'undergraduate research data.xlsx'),
        db_data.get('undergrad_research', [])
    )
    write_advisee_counts(
        os.path.join(service_dir, 'advisee counts.xlsx'),
        db_data.get('advisee_counts', [])
    )
    # Advising evaluation data uses a special university format (STRM codes, specific columns)
    # that cannot be replicated from our DB — skip writing this file so make_far skips it
    # if db_data.get('advising_evals'):
    #     write_advising_evaluation(
    #         os.path.join(service_dir, 'advising evaluation data.xlsx'),
    #         db_data.get('advising_evals', [])
    #     )
    write_teaching_evaluation(
        os.path.join(teach_dir, 'teaching evaluation data.xlsx'),
        db_data.get('teaching', [])
    )
    write_prospective_visits(
        os.path.join(service_dir, 'prospective visit data.xlsx'),
        db_data.get('prospective_visits', [])
    )
