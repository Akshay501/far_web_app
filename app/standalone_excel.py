# app/standalone_excel.py
# FAR Universal (Standalone Mode) — Excel read/write operations.
#
# WHY THIS FILE EXISTS:
# In Clarkson Mode, data lives in MySQL. In Standalone Mode, data lives
# in the professor's Excel files. This file provides the same CRUD
# operations but targeting Excel files directly instead of a database.
#
# PATTERN:
# read_<section>(folder)  → returns list of dicts (like execute_query)
# write_<section>(folder, rows) → writes list of dicts back to Excel
# append_row_<section>(folder, form_data) → adds one new row
# update_row_<section>(folder, row_idx, form_data) → updates one row
# delete_row_<section>(folder, row_idx) → deletes one row

import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime


# ─── Helper functions ──────────────────────────────────────────────────────────

def _get_path(folder, subfolder, filename):
    """Build the full path to an Excel file."""
    return os.path.join(folder, subfolder, filename)


def _read_excel(path):
    """
    Read an Excel file and return a list of dicts.
    Row 1 = headers, Row 2+ = data.
    Returns empty list if file doesn't exist.
    """
    if not os.path.exists(path):
        return []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append(dict(zip(headers, row)))
        return rows
    except Exception:
        return []


def _write_excel(path, headers, rows):
    """
    Write a list of dicts back to an Excel file.
    Creates the file if it doesn't exist.
    Preserves the sheet name 'Data'.
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    wb.save(path)


def _delete_row(path, row_idx):
    """
    Delete row at index row_idx (0-based) from an Excel file.
    Row 0 = first data row (after headers).
    """
    if not os.path.exists(path):
        return
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # +2 because: +1 for 1-based indexing, +1 for header row
    excel_row = row_idx + 2
    ws.delete_rows(excel_row)
    wb.save(path)


def _append_row(path, headers, values):
    """
    Append a new row to an existing Excel file.
    values is a list matching the order of headers.
    """
    if not os.path.exists(path):
        # Create new file with headers
        _write_excel(path, headers, [])
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.append(values)
    wb.save(path)


def _update_row(path, headers, row_idx, values):
    """
    Update an existing row in an Excel file.
    row_idx is 0-based (first data row = 0).
    values is a list matching the order of headers.
    """
    if not os.path.exists(path):
        return
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    excel_row = row_idx + 2  # +1 for 1-based, +1 for header
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=excel_row, column=col_idx).value = value
    wb.save(path)


# ─── Grants ───────────────────────────────────────────────────────────────────

GRANTS_FILE = ('Proposals & Grants', 'grants.xlsx')
GRANTS_HEADERS = [
    'Sponsor', 'Allocated Amt', 'Total Cost', 'Funded?',
    'Title', 'Begin Date', 'End Date', 'Submit Date',
    'Principal Investigators'
]


def read_grants(folder):
    return _read_excel(_get_path(folder, *GRANTS_FILE))


def append_grant(folder, form):
    path = _get_path(folder, *GRANTS_FILE)
    _append_row(path, GRANTS_HEADERS, [
        form.get('sponsor', ''),
        _to_float(form.get('allocated_amt')),
        _to_float(form.get('total_cost')),
        form.get('funded', 'N'),
        form.get('title', ''),
        form.get('begin_date') or None,
        form.get('end_date') or None,
        form.get('submit_date') or None,
        form.get('principal_investigators', ''),
    ])


def update_grant(folder, row_idx, form):
    path = _get_path(folder, *GRANTS_FILE)
    _update_row(path, GRANTS_HEADERS, row_idx, [
        form.get('sponsor', ''),
        _to_float(form.get('allocated_amt')),
        _to_float(form.get('total_cost')),
        form.get('funded', 'N'),
        form.get('title', ''),
        form.get('begin_date') or None,
        form.get('end_date') or None,
        form.get('submit_date') or None,
        form.get('principal_investigators', ''),
    ])


def delete_grant(folder, row_idx):
    _delete_row(_get_path(folder, *GRANTS_FILE), row_idx)


def duplicate_grant(folder, row_idx):
    rows = read_grants(folder)
    if 0 <= row_idx < len(rows):
        row = rows[row_idx]
        path = _get_path(folder, *GRANTS_FILE)
        headers = list(row.keys())
        values = list(row.values())
        _append_row(path, headers, values)


# ─── Proposals ────────────────────────────────────────────────────────────────

PROPOSALS_FILE = ('Proposals & Grants', 'proposals & grants.xlsx')
PROPOSALS_HEADERS = [
    'Sponsor', 'Allocated Amt', 'Total Cost', 'Funded?',
    'Title', 'Begin Date', 'End Date', 'Submit Date',
    'Principal Investigators'
]


def read_proposals(folder):
    return _read_excel(_get_path(folder, *PROPOSALS_FILE))


def append_proposal(folder, form):
    path = _get_path(folder, *PROPOSALS_FILE)
    _append_row(path, PROPOSALS_HEADERS, [
        form.get('sponsor', ''),
        _to_float(form.get('allocated_amt')),
        _to_float(form.get('total_cost')),
        form.get('funded', 'N'),
        form.get('title', ''),
        form.get('begin_date') or None,
        form.get('end_date') or None,
        form.get('submit_date') or None,
        form.get('principal_investigators', ''),
    ])


def update_proposal(folder, row_idx, form):
    path = _get_path(folder, *PROPOSALS_FILE)
    _update_row(path, PROPOSALS_HEADERS, row_idx, [
        form.get('sponsor', ''),
        _to_float(form.get('allocated_amt')),
        _to_float(form.get('total_cost')),
        form.get('funded', 'N'),
        form.get('title', ''),
        form.get('begin_date') or None,
        form.get('end_date') or None,
        form.get('submit_date') or None,
        form.get('principal_investigators', ''),
    ])


def delete_proposal(folder, row_idx):
    _delete_row(_get_path(folder, *PROPOSALS_FILE), row_idx)


def duplicate_proposal(folder, row_idx):
    rows = read_proposals(folder)
    if 0 <= row_idx < len(rows):
        row = rows[row_idx]
        path = _get_path(folder, *PROPOSALS_FILE)
        headers = list(row.keys())
        values = list(row.values())
        _append_row(path, headers, values)


# ─── Service ──────────────────────────────────────────────────────────────────

SERVICE_FILE = ('Service', 'service data.xlsx')
SERVICE_HEADERS = [
    'Description', 'Type', 'Position', 'Term',
    'Calendar Year', 'Hours/Semester', 'Comments'
]


def read_service(folder):
    return _read_excel(_get_path(folder, *SERVICE_FILE))


def append_service(folder, form):
    path = _get_path(folder, *SERVICE_FILE)
    _append_row(path, SERVICE_HEADERS, [
        form.get('description', ''),
        form.get('type', ''),
        form.get('position', ''),
        form.get('term', ''),
        _to_int(form.get('calendar_year')),
        _to_float(form.get('hours_semester')),
        form.get('comments', ''),
    ])


def update_service(folder, row_idx, form):
    path = _get_path(folder, *SERVICE_FILE)
    _update_row(path, SERVICE_HEADERS, row_idx, [
        form.get('description', ''),
        form.get('type', ''),
        form.get('position', ''),
        form.get('term', ''),
        _to_int(form.get('calendar_year')),
        _to_float(form.get('hours_semester')),
        form.get('comments', ''),
    ])


def delete_service(folder, row_idx):
    _delete_row(_get_path(folder, *SERVICE_FILE), row_idx)


def duplicate_service(folder, row_idx):
    rows = read_service(folder)
    if 0 <= row_idx < len(rows):
        row = rows[row_idx]
        path = _get_path(folder, *SERVICE_FILE)
        headers = list(row.keys())
        values = list(row.values())
        _append_row(path, headers, values)


# ─── Awards (Personal) ────────────────────────────────────────────────────────

PERSONAL_AWARDS_FILE = ('Awards', 'personal awards data.xlsx')
PERSONAL_AWARDS_HEADERS = ['Title', 'Type', 'Year']


def read_personal_awards(folder):
    return _read_excel(_get_path(folder, *PERSONAL_AWARDS_FILE))


def append_personal_award(folder, form):
    path = _get_path(folder, *PERSONAL_AWARDS_FILE)
    _append_row(path, PERSONAL_AWARDS_HEADERS, [
        form.get('title', ''),
        form.get('type', ''),
        _to_int(form.get('year')),
    ])


def update_personal_award(folder, row_idx, form):
    path = _get_path(folder, *PERSONAL_AWARDS_FILE)
    _update_row(path, PERSONAL_AWARDS_HEADERS, row_idx, [
        form.get('title', ''),
        form.get('type', ''),
        _to_int(form.get('year')),
    ])


def delete_personal_award(folder, row_idx):
    _delete_row(_get_path(folder, *PERSONAL_AWARDS_FILE), row_idx)


def duplicate_personal_award(folder, row_idx):
    rows = read_personal_awards(folder)
    if 0 <= row_idx < len(rows):
        row = rows[row_idx]
        path = _get_path(folder, *PERSONAL_AWARDS_FILE)
        headers = list(row.keys())
        values = list(row.values())
        _append_row(path, headers, values)


# ─── Awards (Student) ─────────────────────────────────────────────────────────

STUDENT_AWARDS_FILE = ('Awards', 'student awards data.xlsx')
STUDENT_AWARDS_HEADERS = ['Student', 'Title', 'Amount', 'Category', 'Type', 'Year']


def read_student_awards(folder):
    return _read_excel(_get_path(folder, *STUDENT_AWARDS_FILE))


def append_student_award(folder, form):
    path = _get_path(folder, *STUDENT_AWARDS_FILE)
    _append_row(path, STUDENT_AWARDS_HEADERS, [
        form.get('student', ''),
        form.get('title', ''),
        _to_float(form.get('amount')),
        form.get('category', ''),
        form.get('type', ''),
        _to_int(form.get('year')),
    ])


def update_student_award(folder, row_idx, form):
    path = _get_path(folder, *STUDENT_AWARDS_FILE)
    _update_row(path, STUDENT_AWARDS_HEADERS, row_idx, [
        form.get('student', ''),
        form.get('title', ''),
        _to_float(form.get('amount')),
        form.get('category', ''),
        form.get('type', ''),
        _to_int(form.get('year')),
    ])


def delete_student_award(folder, row_idx):
    _delete_row(_get_path(folder, *STUDENT_AWARDS_FILE), row_idx)


def duplicate_student_award(folder, row_idx):
    rows = read_student_awards(folder)
    if 0 <= row_idx < len(rows):
        row = rows[row_idx]
        path = _get_path(folder, *STUDENT_AWARDS_FILE)
        headers = list(row.keys())
        values = list(row.values())
        _append_row(path, headers, values)


# ─── Scholarship (Current Students) ──────────────────────────────────────────

CURRENT_STUDENTS_FILE = ('Scholarship', 'current student data.xlsx')
CURRENT_STUDENTS_HEADERS = ['Student Name', 'Current Program', 'Start Date']


def read_current_students(folder):
    return _read_excel(_get_path(folder, *CURRENT_STUDENTS_FILE))


def append_current_student(folder, form):
    path = _get_path(folder, *CURRENT_STUDENTS_FILE)
    _append_row(path, CURRENT_STUDENTS_HEADERS, [
        form.get('student_name', ''),
        form.get('current_program', ''),
        form.get('start_date') or None,
    ])


def update_current_student(folder, row_idx, form):
    path = _get_path(folder, *CURRENT_STUDENTS_FILE)
    _update_row(path, CURRENT_STUDENTS_HEADERS, row_idx, [
        form.get('student_name', ''),
        form.get('current_program', ''),
        form.get('start_date') or None,
    ])


def delete_current_student(folder, row_idx):
    _delete_row(_get_path(folder, *CURRENT_STUDENTS_FILE), row_idx)


def duplicate_current_student(folder, row_idx):
    rows = read_current_students(folder)
    if 0 <= row_idx < len(rows):
        row = rows[row_idx]
        path = _get_path(folder, *CURRENT_STUDENTS_FILE)
        headers = list(row.keys())
        values = list(row.values())
        _append_row(path, headers, values)


# ─── Scholarship (Thesis) ─────────────────────────────────────────────────────

THESIS_FILE = ('Scholarship', 'thesis data.xlsx')
THESIS_HEADERS = ['Student', 'Start Date', 'Year', 'Degree', 'Advisor', 'Title', 'Comments']


def read_thesis(folder):
    return _read_excel(_get_path(folder, *THESIS_FILE))


def append_thesis(folder, form):
    path = _get_path(folder, *THESIS_FILE)
    _append_row(path, THESIS_HEADERS, [
        form.get('student', ''),
        _to_int(form.get('start_date_year')) or 0,
        _to_int(form.get('year')),
        form.get('degree', ''),
        form.get('advisor', ''),
        form.get('title', ''),
        form.get('comments', ''),
    ])


def update_thesis(folder, row_idx, form):
    path = _get_path(folder, *THESIS_FILE)
    _update_row(path, THESIS_HEADERS, row_idx, [
        form.get('student', ''),
        _to_int(form.get('start_date_year')) or 0,
        _to_int(form.get('year')),
        form.get('degree', ''),
        form.get('advisor', ''),
        form.get('title', ''),
        form.get('comments', ''),
    ])


def delete_thesis(folder, row_idx):
    _delete_row(_get_path(folder, *THESIS_FILE), row_idx)


def duplicate_thesis(folder, row_idx):
    rows = read_thesis(folder)
    if 0 <= row_idx < len(rows):
        row = rows[row_idx]
        path = _get_path(folder, *THESIS_FILE)
        headers = list(row.keys())
        values = list(row.values())
        _append_row(path, headers, values)


# ─── Teaching (Read-only) ─────────────────────────────────────────────────────

TEACHING_FILE = ('Teaching', 'teaching evaluation data.xlsx')


def read_teaching(folder):
    return _read_excel(_get_path(folder, *TEACHING_FILE))


# ─── File inventory ───────────────────────────────────────────────────────────

def get_file_inventory(folder):
    """
    Check which Excel files exist in the uploaded folder.
    Returns a dict of section → (exists, path).
    Used on the dashboard to show the professor what data is available.
    """
    files = {
        'grants':            ('Proposals & Grants', 'grants.xlsx'),
        'proposals':         ('Proposals & Grants', 'proposals & grants.xlsx'),
        'expenditures':      ('Proposals & Grants', 'expenditures.xlsx'),
        'personal_awards':   ('Awards', 'personal awards data.xlsx'),
        'student_awards':    ('Awards', 'student awards data.xlsx'),
        'current_students':  ('Scholarship', 'current student data.xlsx'),
        'thesis':            ('Scholarship', 'thesis data.xlsx'),
        'service':           ('Service', 'service data.xlsx'),
        'reviews':           ('Service', 'reviews data.xlsx'),
        'prof_development':  ('Service', 'professional development data.xlsx'),
        'undergrad_research':('Service', 'undergraduate research data.xlsx'),
        'teaching':          ('Teaching', 'teaching evaluation data.xlsx'),
        'make_cv_far':       ('make_cv', 'FAR'),
    }
    inventory = {}
    for key, (subfolder, filename) in files.items():
        path = os.path.join(folder, subfolder, filename)
        inventory[key] = {
            'exists': os.path.exists(path),
            'path': path,
            'is_dir': os.path.isdir(path),
        }
    return inventory


# ─── Type conversion helpers ──────────────────────────────────────────────────

def _to_float(val):
    """Safely convert form string to float."""
    try:
        return float(val) if val else None
    except (ValueError, TypeError):
        return None


def _to_int(val):
    """Safely convert form string to int."""
    try:
        return int(val) if val else None
    except (ValueError, TypeError):
        return None


# ─── Publications (.bib) ──────────────────────────────────────────────────────

BIB_FILE = ('Scholarship', 'scholarship.bib')


def get_bib_path(folder):
    return _get_path(folder, *BIB_FILE)


def bib_exists(folder):
    return os.path.exists(get_bib_path(folder))
