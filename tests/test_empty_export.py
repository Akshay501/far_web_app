"""
Tests for empty-table export semantics (Issue #13).

make_far reads three summary sheets positionally (.iloc[-1] on the last
row): advisee counts, expenditures, and prospective visits. A sheet that
exists with zero data rows crashes generation ("single positional
indexer is out-of-bounds"); an ABSENT sheet makes make_far skip the
section — its designed no-data behavior (other make_far sections delete
their own output when they find zero rows).

The contract: export_all writes these three sheets only when rows
exist, and REMOVES a stale sheet when the table has gone empty (so a
professor who deletes their last row doesn't keep exporting old
numbers). All other sheets keep writing even when empty — proven
harmless by the professor-10 crash log, where every iterated section
processed zero-row sheets fine before the advisee-counts crash.

No database involved: export_all takes a plain db_data dict.
"""
import pytest
from openpyxl import load_workbook

from app.excel_export import export_all

# (subfolder, filename) for the three positionally-read summary sheets
TRIO = [
    ('Service', 'advisee counts.xlsx'),
    ('Proposals & Grants', 'expenditures.xlsx'),
    ('Service', 'prospective visit data.xlsx'),
]


@pytest.fixture
def prof_folder(tmp_path):
    """A bare professor folder — export_all's ensure() creates subdirs."""
    return tmp_path / 'prof'


def _path(folder, sub, name):
    return folder / sub / name


def test_empty_tables_do_not_create_summary_sheets(prof_folder):
    export_all(9001, str(prof_folder), {})

    for sub, name in TRIO:
        assert not _path(prof_folder, sub, name).exists(), \
            f'{name} must be absent when the table has zero rows'


def test_other_sheets_still_written_when_empty(prof_folder):
    """Empty is harmless for the iterated sheets — they keep existing so
    their DB truth (including 'now empty') always overwrites the disk."""
    export_all(9001, str(prof_folder), {})

    assert _path(prof_folder, 'Awards', 'personal awards data.xlsx').exists()
    assert _path(prof_folder, 'Teaching',
                 'teaching evaluation data.xlsx').exists()
    assert _path(prof_folder, 'Service', 'service data.xlsx').exists()


def test_stale_summary_sheets_are_removed(prof_folder):
    """A professor who once had data (sheets on disk) and now has none:
    the stale sheets must be deleted, not left to feed old numbers into
    the FAR."""
    for sub, name in TRIO:
        p = _path(prof_folder, sub, name)
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text('stale content from a previous export')

    export_all(9001, str(prof_folder), {})

    for sub, name in TRIO:
        assert not _path(prof_folder, sub, name).exists(), \
            f'stale {name} must be removed when the table is empty'


def test_summary_sheets_written_when_rows_exist(prof_folder):
    db_data = {
        'advisee_counts': [
            {'Advisor Name': 'T. Prof', 'Advisee Count': 7, 'Year': 2025},
        ],
        'expenditures': [
            {'Year': 2025, 'Name': 'T. Prof', 'Expenditure': 100.0,
             'Indirect': 10.0, 'Tuition': 5.0, 'Salary Recovery': 1.0},
        ],
        'prospective_visits': [
            {'Year': 2025, 'Staff': 'T. Prof', 'Visits': 3, 'Deposits': 1},
        ],
    }

    export_all(9001, str(prof_folder), db_data)

    for sub, name in TRIO:
        p = _path(prof_folder, sub, name)
        assert p.exists()
        ws = load_workbook(p).active
        assert ws.max_row >= 2, f'{name} should have header + data rows'
