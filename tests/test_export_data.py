"""
Tests for professor data export (Issue #7) — revised design.

The contract:

  POST /professor/export
      -> a zip of the logged-in professor's own folder, REFRESHED from
         the database first (bib + Excel files), so the export carries
         current data even for a professor who has never generated.
         POST because it writes files to disk.
  POST /admin/professor/<pk>/export   (admin only)
      -> the same for any professor; non-admins are bounced.

  Contents: the whole folder MINUS LaTeX build artifacts, so the export
  is a working folder someone can drop on a laptop and run make_cv in.
  A missing folder is healed first — an export should never fail with
  "no folder", it should make one.

  Filename: far_data_<last>_<first>_<date>.zip via safe_slug(), so a
  download identifies its professor once detached.

The content test is the heart of this file: it seeds a real award in
the DATABASE (touching no files) and asserts its title appears inside
the xlsx inside the zip. Filename-only assertions would pass on a zip
of empty skeletons — asserting structure instead of content is how a
test lies.
"""
import io
import zipfile

import pytest
from openpyxl import load_workbook

from app.utils import execute_query, safe_slug

AWARD_TITLE = 'Export Content Probe Award ZYXW'


# ------------------------------------------------------------------- slug

@pytest.mark.parametrize('raw,expected', [
    ('Thugudam', 'thugudam'),
    ('Van Der Berg', 'van_der_berg'),
    ("O'Brien", 'obrien'),
    ('José García', 'jose_garcia'),
    ('Smith-Jones', 'smith-jones'),
    ('  padded  ', 'padded'),
    ('../../etc/passwd', 'etcpasswd'),
    ('', 'unknown'),
])
def test_safe_slug(raw, expected):
    """Filesystem- and header-safe: lowercase ascii, no spaces, no path
    or quote characters — and never empty."""
    assert safe_slug(raw) == expected


# ---------------------------------------------------------------- fixtures

@pytest.fixture
def prof_folder(scaffold, seed_professor):
    """A professor folder holding data files plus the build artifacts a
    real generation leaves behind."""
    folder = scaffold['root'] / str(seed_professor['professor_key'])
    (folder / 'Awards').mkdir(parents=True)
    (folder / 'Scholarship').mkdir(parents=True)
    (folder / 'make_cv' / 'FAR' / 'Tables_far').mkdir(parents=True)

    (folder / 'Scholarship' / 'scholarship.bib').write_text('% bib\n')
    (folder / '.scaffold_version').write_text('commit = abc123\n')
    (folder / 'make_cv' / 'FAR' / 'far.tex').write_text('% tex\n')
    (folder / 'make_cv' / 'FAR' / 'far.pdf').write_bytes(b'%PDF-fake')

    # Build artifacts — regenerated every run, must not be exported.
    for junk in ('far.aux', 'far.log', 'far.out', 'far.toc', 'far.bbl',
                 'far.bcf', 'far.blg', 'far.run.xml'):
        (folder / 'make_cv' / 'FAR' / junk).write_text('junk')
    (folder / 'make_cv' / 'FAR' / 'Tables_far' / 'journal.tex').write_text('t')

    return folder


@pytest.fixture
def seeded_award(app, seed_professor):
    """One student award in the DATABASE only — no files touched. The
    export's refresh step is what must carry it onto disk."""
    with app.app_context():
        key = execute_query(
            'INSERT INTO AWARDS (Title, Year, `Award Type`) '
            'VALUES (%s, %s, %s)',
            (AWARD_TITLE, 2025, 'Graduate'), commit=True, lastrowid=True)
        execute_query(
            'INSERT INTO STUDENTAWARDS '
            '(`Award Key`, ProfessorKey, Student, Amount, Category) '
            'VALUES (%s, %s, %s, %s, %s)',
            (key, seed_professor['professor_key'], 'Probe Student',
             100, 'Research'), commit=True)
    yield AWARD_TITLE
    with app.app_context():
        execute_query('DELETE FROM STUDENTAWARDS WHERE `Award Key` = %s',
                      (key,), commit=True)
        execute_query('DELETE FROM AWARDS WHERE `Award Key` = %s',
                      (key,), commit=True)


def _zip(resp):
    return zipfile.ZipFile(io.BytesIO(resp.data))


# ------------------------------------------------------- professor export

def test_professor_exports_own_folder(logged_in_client, seed_professor,
                                      prof_folder):
    resp = logged_in_client.post('/professor/export')

    assert resp.status_code == 200
    assert resp.mimetype == 'application/zip'

    names = _zip(resp).namelist()
    assert any(n.endswith('scholarship.bib') for n in names)
    assert any(n.endswith('far.tex') for n in names)
    assert any(n.endswith('.scaffold_version') for n in names)
    # Written by the refresh even though the fixture didn't create it:
    assert any(n.endswith('personal awards data.xlsx') for n in names)


def test_export_contains_current_db_data(logged_in_client, seed_professor,
                                         scaffold, seeded_award):
    """THE core test. Data lives only in the DB; no folder even exists.
    The zip must contain the award title inside the student-awards xlsx
    — proof that export refreshes from the database rather than
    shipping whatever stale files happen to be on disk."""
    resp = logged_in_client.post('/professor/export')
    assert resp.status_code == 200

    zf = _zip(resp)
    name = next(n for n in zf.namelist()
                if n.endswith('student awards data.xlsx'))
    ws = load_workbook(io.BytesIO(zf.read(name))).active
    cells = [str(c) for row in ws.iter_rows(values_only=True)
             for c in row if c is not None]
    assert AWARD_TITLE in cells, \
        'the DB award must reach the exported spreadsheet'


def test_export_excludes_build_artifacts(logged_in_client, seed_professor,
                                         prof_folder):
    names = _zip(logged_in_client.post('/professor/export')).namelist()

    for junk in ('far.aux', 'far.log', 'far.out', 'far.toc', 'far.bbl',
                 'far.bcf', 'far.blg', 'far.run.xml'):
        assert not any(n.endswith(junk) for n in names), \
            f'{junk} is a build artifact and must not be exported'
    assert not any('Tables_far' in n for n in names)


def test_export_filename_identifies_the_professor(logged_in_client,
                                                  seed_professor,
                                                  prof_folder):
    resp = logged_in_client.post('/professor/export')
    disposition = resp.headers.get('Content-Disposition', '')

    assert 'far_data_' in disposition
    assert 'professor' in disposition.lower()   # seeded LastName, slugged
    assert '.zip' in disposition


def test_export_heals_a_missing_folder(logged_in_client, seed_professor,
                                       scaffold):
    """No folder on disk: the export builds one instead of failing."""
    folder = scaffold['root'] / str(seed_professor['professor_key'])
    assert not folder.exists()

    resp = logged_in_client.post('/professor/export')

    assert resp.status_code == 200
    assert folder.is_dir()
    assert any(n.endswith('.scaffold_version')
               for n in _zip(resp).namelist())


# ----------------------------------------------------------- admin export

def test_admin_exports_any_professor(admin_client, seed_professor,
                                     prof_folder):
    resp = admin_client.post(
        f"/admin/professor/{seed_professor['professor_key']}/export")

    assert resp.status_code == 200
    assert resp.mimetype == 'application/zip'
    assert any(n.endswith('scholarship.bib')
               for n in _zip(resp).namelist())


def test_non_admin_cannot_use_admin_export(logged_in_client,
                                           seed_professor, prof_folder):
    resp = logged_in_client.post(
        f"/admin/professor/{seed_professor['professor_key']}/export")

    assert resp.status_code == 302
    assert '/login' in resp.headers['Location']
