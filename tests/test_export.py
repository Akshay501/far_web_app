"""
Tests for excel_export.export_all — the seam between the database and make_cv.

These guard the *contract*: each make_cv converter opens a specific file,
reads a specific sheet, and accesses specific columns by name. If any of
those drift, the matching section silently vanishes from the generated
FAR (no error — just a missing section). These tests turn that silent
failure into a loud one.

Column requirements below were read directly from the make_cv 1.0.3
converters, e.g. service2latex_far.py accesses df.loc[count,"Term"], so
"Term" is required in the service file.

No database, no login, no LaTeX — export_all is a pure function, so these
run in milliseconds against a throwaway temp directory.

Place this file at: tests/test_export.py
"""
import os
from datetime import date

import openpyxl
import pytest

from app.excel_export import export_all


# ---------------------------------------------------------------------------
# The contract, one row per file:
#   (path relative to professor folder, required sheet, required columns)
# sheet == None means the converter reads the first sheet positionally,
# so the sheet's name does not matter.
# ---------------------------------------------------------------------------
CONTRACT = [
    # Verified from personal_awards2latex_far.py line 44
    ('Awards/personal awards data.xlsx', 'Data',
     ['Year', 'Type', 'Title']),

    # Verified from student_awards2latex_far.py line 41
    ('Awards/student awards data.xlsx', 'Data',
     ['Year', 'Title', 'Student']),

    # Verified from service2latex_far.py line 40
    ('Service/service data.xlsx', 'Data',
     ['Calendar Year', 'Term', 'Type', 'Description', 'Hours/Semester']),

    # Verified from reviews2latex_far.py lines 23-40
    ('Service/reviews data.xlsx', 'Data',
     ['Journal', 'Start', 'Rounds']),

    # Verified from UR2latex_far.py line 40
    ('Service/undergraduate research data.xlsx', 'Data',
     ['Students', 'Title', 'Program Type', 'Term', 'Calendar Year']),

    # Verified from thesis2latex_far.py line 204 (the completed-thesis table).
    # read_thesis_file may reference more columns; not asserted here.
    ('Scholarship/thesis data.xlsx', 'Data',
     ['Student', 'Title', 'Year', 'Degree']),

    # Verified from thesis2latex_far.py line 146 (sort keys) and 189/191.
    # NOTE: "Title" is deliberately NOT required -- line 147 defaults it to
    # "--" when absent. Supplying it would show each in-progress student's
    # research title instead of "--", but its absence is not an error.
    ('Scholarship/current student data.xlsx', 'Data',
     ['Student Name', 'Current Program', 'Start Date']),

    # Verified from props2latex_far.py lines 21, 71-72
    ('Proposals & Grants/proposals & grants.xlsx', 'Data',
     ['Sponsor', 'Title', 'Allocated Amt', 'Total Cost',
      'Begin Date', 'End Date', 'Submit Date']),

    # Verified from grants2latex_far.py lines 78-79
    ('Proposals & Grants/grants.xlsx', 'Data',
     ['Sponsor', 'Title', 'Allocated Amt', 'Total Cost',
      'Begin Date', 'End Date']),

    # make_far.py lines 98-103 reads this with pd.read_excel(filename,
    # skiprows=0) -- no sheet_name, so it takes the first sheet whatever
    # it is called, and reads the LAST row.
    ('Proposals & Grants/expenditures.xlsx', None,
     ['Year', 'Name', 'Expenditure', 'Indirect', 'Tuition',
      'Salary Recovery']),

    # Verified from teaching2latex_far.py lines 64, 81-83, 86.
    # 'component', 'course_title' and 'STRM' are deliberately NOT required:
    # lines 66-77 default component to "LEC", course_title to "" and derive
    # STRM from term when they are absent.
    ('Teaching/teaching evaluation data.xlsx', 'Data',
     ['term', 'combined_course_num', 'combined_num_sec', 'enrollment',
      'count_19', 'mean_19', 'count_20', 'mean_20']),

    # No FAR converter reads this file directly; sheet name only.
    ('Service/professional development data.xlsx', 'Data', []),
]

# Folders make_cv expects inside a professor folder.
EXPECTED_FOLDERS = ['Awards', 'Proposals & Grants', 'Scholarship',
                    'Service', 'Teaching']


@pytest.fixture
def exported(tmp_path):
    """Run export_all into a throwaway folder with a little known data."""
    db_data = {
        'reviews': [
            {'Journal': 'Journal of Testing', 'Start Date': date(2025, 1, 15),
             'Rounds': 3},
        ],
        'expenditures': [
            {'Year': 2025, 'Name': 'Test Grant', 'Expenditure': 1000,
             'Indirect': 200, 'Tuition': 300, 'Salary Recovery': 400},
        ],
    }
    export_all(9001, str(tmp_path), db_data)
    return tmp_path


def _headers(path, sheet):
    """Return the header row of the given sheet (or the first sheet)."""
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    return [c.value for c in ws[1]], wb.sheetnames


# ---------------------------------------------------------------------------
# One test per file, generated by parametrize. Each shows up separately in
# the pytest output, so a failure names the exact file at fault.
# ---------------------------------------------------------------------------

@pytest.mark.parametrize('relpath,sheet,columns', CONTRACT,
                         ids=[c[0] for c in CONTRACT])
def test_file_is_created(exported, relpath, sheet, columns):
    """make_cv looks for this exact filename in this exact folder."""
    assert os.path.isfile(os.path.join(str(exported), relpath)), \
        f'export_all did not write {relpath}'


@pytest.mark.parametrize('relpath,sheet,columns', CONTRACT,
                         ids=[c[0] for c in CONTRACT])
def test_sheet_name_matches_converter(exported, relpath, sheet, columns):
    """
    Converters call read_excel(sheet_name='Data'). A different sheet name
    makes the read fail and the section disappears from the FAR.
    """
    if sheet is None:
        pytest.skip('converter reads the first sheet positionally')
    _, sheetnames = _headers(os.path.join(str(exported), relpath), None)
    assert sheet in sheetnames, \
        f'{relpath} has sheets {sheetnames}, converter needs {sheet!r}'


@pytest.mark.parametrize('relpath,sheet,columns', CONTRACT,
                         ids=[c[0] for c in CONTRACT])
def test_columns_match_converter(exported, relpath, sheet, columns):
    """Every column the converter accesses by name must be present."""
    if not columns:
        pytest.skip('column requirements not yet pinned down for this file')
    headers, _ = _headers(os.path.join(str(exported), relpath), sheet)
    missing = [c for c in columns if c not in headers]
    assert not missing, \
        f'{relpath} is missing {missing}; it has {headers}'


# ---------------------------------------------------------------------------
# Structure and data checks
# ---------------------------------------------------------------------------

@pytest.mark.parametrize('folder', EXPECTED_FOLDERS)
def test_expected_subfolder_exists(exported, folder):
    """make_cv expects this exact folder layout inside a professor folder."""
    assert os.path.isdir(os.path.join(str(exported), folder))


def test_reviews_data_row_is_written(exported):
    """Headers alone aren't enough — the seeded review must appear as a row."""
    wb = openpyxl.load_workbook(
        os.path.join(str(exported), 'Service/reviews data.xlsx'))
    values = [c.value for c in wb['Data'][2]]
    assert 'Journal of Testing' in values
    assert 3 in values


def test_expenditures_data_row_is_written(exported):
    """make_far reads the LAST row of this file, so a row must exist."""
    wb = openpyxl.load_workbook(
        os.path.join(str(exported), 'Proposals & Grants/expenditures.xlsx'))
    ws = wb[wb.sheetnames[0]]
    values = [c.value for c in ws[2]]
    assert 'Test Grant' in values
